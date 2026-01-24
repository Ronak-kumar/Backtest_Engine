import pandas as pd
import os
import warnings
import calendar
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

def heat_map(filepath, indices):

    full_path = filepath
    path = filepath.parent
    df = pd.read_csv(full_path)
    df.rename(columns={"PnL": "PNL"}, inplace=True)
    df['Timestamp'] = pd.to_datetime(df['Timestamp'])

    monthly_pnl = df.groupby(df['Timestamp'].dt.to_period('M'))['Equity Return'].sum()
    monthly_pnl_df = monthly_pnl.reset_index()

    monthly_pnl_df['Year'] = monthly_pnl_df['Timestamp'].apply(lambda x: x.year)
    monthly_pnl_df['Month'] = monthly_pnl_df['Timestamp'].apply(lambda x: x.month)

    monthly_pnl_df['Month'] = monthly_pnl_df['Month'].apply(lambda x: calendar.month_abbr[x])
    monthly_pnl_df = monthly_pnl_df[['Timestamp', 'Equity Return', 'Year', 'Month']]

    month_order = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    monthly_pnl_df["Month"] = pd.Categorical(
        monthly_pnl_df["Month"], 
        categories=month_order, 
        ordered=True
    )

    monthly_pnl_df = monthly_pnl_df.sort_values(["Year", "Month"])
    months = monthly_pnl_df[["Year", "Month"]].drop_duplicates().apply(
        lambda x: f"{x['Year']} {x['Month']}", axis=1
    ).tolist()
    #months = monthly_pnl_df['Month'].tolist()
    months.append('Total')
    x = monthly_pnl_df.groupby(['Year', 'Month'])['Equity Return'].mean()

    broad_returns = x.unstack()
    broad_returns['Total'] = broad_returns[broad_returns.columns].sum(axis=1)
    broad_returns.fillna(0, inplace = True)

    grand_total = pd.Series({ 'Total': broad_returns['Total'].sum() }, name='Sum')

    # Append it
    broad_returns = pd.concat([broad_returns, grand_total.to_frame().T])

    # Reset index so Year/Sum becomes a column
    broad_returns = broad_returns.reset_index()

    # Rename the index column to empty string
    broad_returns.rename(columns={'index': ''}, inplace=True)

    # Replace "Sum" with empty in that column
    broad_returns.loc[broad_returns[''] == 'Sum', ''] = ''
    print(broad_returns)
    

    current_dd = df['Drawdown'].iloc[-1]
    max_drawdown = df['Drawdown'].max()

    Second_table = pd.DataFrame({
        "Metric": ["Current_DD", "Max_DD"],
        "Value": [current_dd, max_drawdown]
    })

    drawdowns = df.sort_values(['Drawdown', 'Timestamp'], ascending=False)
    drawdowns = drawdowns.loc[drawdowns['Drawdown'] > 0, ['Timestamp', 'Drawdown']]
    drawdowns.rename(columns = {'Timestamp' : "Date"}, inplace = True)
    drawdowns['Date'] = drawdowns['Date'].dt.date


    latest_date = df['Timestamp'].max().date()

    # make note as a tiny DataFrame (one column)
    note_df = pd.DataFrame({f"Updated till: {latest_date}": []})
    note_df_clean = note_df.reset_index(drop=True)
    broad_returns_clean = broad_returns.reset_index(drop=True)
    drawdowns_clean = drawdowns.reset_index(drop=True)   
    

    output_file = path /f"{indices}_Result_Heatmap.xlsx"
    row_offset = 2  

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        note_df_clean.to_excel(writer, sheet_name="Report", startcol=0, startrow=row_offset, index=False)

        br_startcol = note_df_clean.shape[1] + 1
        broad_returns_clean.to_excel(writer, sheet_name="Report", startcol=br_startcol, startrow=row_offset, index=False)

        st_startcol = br_startcol + broad_returns_clean.shape[1] + 1
        Second_table.to_excel(writer, sheet_name="Report", startcol=st_startcol, startrow=row_offset, index=False)

        dd_startcol = st_startcol + Second_table.shape[1] + 1
        drawdowns_clean.to_excel(writer, sheet_name="Report", startcol=dd_startcol, startrow=row_offset, index=False)

    
    wb = load_workbook(output_file)
    ws = wb["Report"]

    # Border style
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Apply borders only inside one table block
    def apply_table_borders(start_row, start_col, nrows, ncols):
        for r in range(start_row+1, start_row+1+nrows):   # include header row
            for c in range(start_col+1, start_col+1+ncols):   # Excel is 1-based
                ws.cell(row=r, column=c).border = thin_border

    # Borders
    apply_table_borders(row_offset, 0, len(note_df_clean)+1, note_df_clean.shape[1])
    apply_table_borders(row_offset, br_startcol, len(broad_returns_clean)+1, broad_returns_clean.shape[1])
    apply_table_borders(row_offset, st_startcol, len(Second_table)+1, Second_table.shape[1])
    apply_table_borders(row_offset, dd_startcol, len(drawdowns_clean)+1, drawdowns_clean.shape[1])

    
    red_fill   = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")   # light red
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")   # light green

    red_font   = Font(color="9C0006")   # dark red text
    green_font = Font(color="006100")   # dark green text

    # Range for broad_returns values (skip first column = Year)
    first_val_col = br_startcol + 2
    last_val_col  = br_startcol + broad_returns_clean.shape[1]

    cell_range = (
        f"{ws.cell(row=row_offset+2, column=first_val_col).coordinate}:"
        f"{ws.cell(row=row_offset+len(broad_returns_clean)+1, column=last_val_col).coordinate}"
    )

    # Apply conditional formatting
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="lessThan", formula=["0"], fill=red_fill, font=red_font)
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="greaterThan", formula=["0"], fill=green_fill, font=green_font)
    )

    
    br_first_col = br_startcol + 1
    br_last_col = br_startcol + broad_returns_clean.shape[1]

    for col_idx in range(br_first_col, br_last_col + 1):
        col_letter = get_column_letter(col_idx)
        if col_idx == br_first_col:
            ws.column_dimensions[col_letter].width = 10
        else:
            ws.column_dimensions[col_letter].width = 7

    

    def auto_adjust_col_width(ws, df, startcol, row_offset):
        """
        Auto-adjust column widths for a dataframe written into Excel.
        startcol = pandas-style starting column (0-based)
        row_offset = starting row offset used in to_excel
        """
        for i, col_name in enumerate(df.columns):
            col_idx = startcol + i + 1   # Excel is 1-based
            col_letter = get_column_letter(col_idx)

            # Get max length across header + column values
            max_len = max(
                [len(str(col_name))] +
                [len(str(s)) for s in df.iloc[:, i].dropna().astype(str)]
            )

            # Set width with small padding
            ws.column_dimensions[col_letter].width = max_len + 2

    
    auto_adjust_col_width(ws, note_df_clean, 0, row_offset)
    auto_adjust_col_width(ws, Second_table, st_startcol, row_offset)
    auto_adjust_col_width(ws, drawdowns_clean, dd_startcol, row_offset)

    # Save once at the end
    wb.save(output_file)




