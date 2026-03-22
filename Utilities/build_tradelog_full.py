"""
NIFTY Options Strategy — Full Trade Log & Statistical Report
============================================================
Usage:
    python build_tradelog_full.py

Requirements:
    pip install pandas openpyxl numpy

Input:  order_book.csv  (same folder, or change INPUT_PATH below)
Output: NIFTY_Trade_Logs.xlsx
"""

import re, itertools
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── CONFIG ───────────────────────────────────────────────────────────────────
INPUT_PATH  = "order_book.csv"          # ← change if needed
OUTPUT_PATH = "NIFTY_Trade_Logs.xlsx"

# ─── PALETTE ──────────────────────────────────────────────────────────────────
NAVY    = "1B1F2E"
HEADER  = "1E3A5F"
GREEN   = "ECFDF0"
RED     = "FEF2F2"
ALT     = "F0F4FF"
WHITE   = "FFFFFF"
LGRAY   = "F8FAFF"
DGRAY   = "444444"

# ─── STYLE HELPERS ────────────────────────────────────────────────────────────
def fill(h):       return PatternFill("solid", start_color=h, fgColor=h)
def thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

C = Alignment(horizontal="center", vertical="center")
L = Alignment(horizontal="left",   vertical="center")
R = Alignment(horizontal="right",  vertical="center")

def font(sz=9, bold=False, color="000000", name="Arial"):
    return Font(name=name, size=sz, bold=bold, color=color)

def pnl_font(v, sz=9, bold=False):
    c = "1A7A3C" if v > 0 else ("C0392B" if v < 0 else "000000")
    return font(sz=sz, bold=bold, color=c)

def wcell(ws, r, c, val, bg=WHITE, fnt=None, fmt=None, align=C, border=True):
    cell           = ws.cell(r, c)
    cell.value     = val
    cell.fill      = fill(bg)
    cell.font      = fnt or font()
    cell.alignment = align
    if border: cell.border = thin()
    if fmt:    cell.number_format = fmt
    return cell

def header_row(ws, row, labels, bg=HEADER, fg="FFFFFF", sz=10, height=20):
    for c, h in enumerate(labels, 1):
        wcell(ws, row, c, h, bg=bg, fnt=font(sz=sz, bold=True, color=fg))
    ws.row_dimensions[row].height = height

def col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def title_row(ws, text, ncols, row=1, bg=NAVY, sz=13, height=30):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1)
    c.value     = text
    c.font      = font(sz=sz, bold=True, color="FFFFFF")
    c.fill      = fill(bg)
    c.alignment = C
    ws.row_dimensions[row].height = height

def totals_row(ws, r, ncols, formulas: dict):
    for c in range(1, ncols+1):
        cell = ws.cell(r, c)
        cell.fill   = fill(NAVY)
        cell.font   = font(sz=9, bold=True, color="FFFFFF")
        cell.alignment = C
        cell.border = thin()
    ws.cell(r, 1).value = "TOTAL / AVG"
    for col, formula in formulas.items():
        ws.cell(r, col).value          = formula
        ws.cell(r, col).number_format  = "#,##0.00"
    ws.row_dimensions[r].height = 18

# ─── DATA PIPELINE ────────────────────────────────────────────────────────────
def build_trades(path):
    df = pd.read_csv(path)
    df["Timestamp"] = pd.to_datetime(df["Timestamp"])

    def parse(s):
        s = str(s)
        if "Executed entry" in s:
            sl  = re.search(r"sl ([\d.]+)", s)
            tgt = re.search(r"target ([\d.]+)", s)
            st  = re.search(r"strike (\d+)", s)
            return {"event":"ENTRY",
                    "sl":    float(sl.group(1))  if sl  else None,
                    "target":float(tgt.group(1)) if tgt else None,
                    "strike":int(st.group(1))    if st  else None}
        elif "SL Triggered"     in s: return {"event":"SL_EXIT",    "sl":None,"target":None,"strike":None}
        elif "Target Triggered" in s: return {"event":"TARGET_EXIT","sl":None,"target":None,"strike":None}
        elif "Exit time"        in s: return {"event":"TIME_EXIT",  "sl":None,"target":None,"strike":None}
        return {"event":"OTHER","sl":None,"target":None,"strike":None}

    parsed = df["Summary"].apply(parse).apply(pd.Series)
    df = pd.concat([df, parsed], axis=1)
    df["leg"] = df["Summary"].apply(
        lambda s: (re.search(r"leg_[\w\d]+", str(s)) or
                   type("",(),{"group":lambda _,x:None})()).group(0))
    df["option_type"] = df["Ticker"].str.extract(r"(CE|PE)$")

    entries = df[df["event"]=="ENTRY"][
        ["Timestamp","Ticker","Price","leg","option_type","sl","target","strike"]].copy()
    entries.columns = ["entry_ts","Ticker","entry_price","leg","option_type","sl","target","strike"]
    exits = df[df["event"].isin(["SL_EXIT","TARGET_EXIT","TIME_EXIT"])][
        ["Timestamp","Ticker","Price","leg","event"]].copy()
    exits.columns = ["exit_ts","Ticker","exit_price","leg","exit_type"]

    t = pd.merge(entries, exits, on=["Ticker","leg"])
    t = t[t["exit_ts"] >= t["entry_ts"]]
    t = t.sort_values("exit_ts").groupby(["Ticker","leg","entry_ts"]).first().reset_index()

    t["pnl"]          = (t["entry_price"] - t["exit_price"]).round(2)
    t["pnl_pct"]      = (t["pnl"] / t["entry_price"] * 100).round(2)
    t["duration_min"] = ((t["exit_ts"] - t["entry_ts"]).dt.total_seconds()/60).round(0).astype(int)
    t["trade_date"]   = t["entry_ts"].dt.date
    t["entry_time"]   = t["entry_ts"].dt.strftime("%H:%M")
    t["exit_time"]    = t["exit_ts"].dt.strftime("%H:%M")
    t["result"]       = t["exit_type"].map({"SL_EXIT":"SL","TARGET_EXIT":"TARGET","TIME_EXIT":"TIME"})
    t["entry_hour"]   = t["entry_ts"].dt.hour
    t["dow"]          = t["entry_ts"].dt.dayofweek
    t["dow_name"]     = t["entry_ts"].dt.day_name()
    t["year"]         = t["entry_ts"].dt.year
    t["ym"]           = t["entry_ts"].dt.to_period("M").astype(str)
    t["implied_atm"]  = (t["entry_price"] / 0.90).round(2)
    t["dur_bucket"]   = pd.cut(t["duration_min"],
        bins=[0,15,30,60,120,240,360,1440,99999],
        labels=["<15m","15-30m","30-60m","1-2h","2-4h","4-6h","6-24h",">1day"])
    t["ep_bucket"]    = pd.cut(t["entry_price"],
        bins=[0,30,50,75,100,150,200,999],
        labels=["<30","30-50","50-75","75-100","100-150","150-200",">200"])
    return t.sort_values(["trade_date","entry_time"])


def build_daily(t):
    d = t.groupby("trade_date").agg(
        Trades       =("pnl","count"),
        Winners      =("pnl", lambda x:(x>0).sum()),
        Losers       =("pnl", lambda x:(x<0).sum()),
        Gross_PnL    =("pnl","sum"),
        Best_Trade   =("pnl","max"),
        Worst_Trade  =("pnl","min"),
        SL_Count     =("result", lambda x:(x=="SL").sum()),
        Target_Count =("result", lambda x:(x=="TARGET").sum()),
        Time_Count   =("result", lambda x:(x=="TIME").sum()),
        Avg_Duration =("duration_min","mean"),
    ).reset_index().sort_values("trade_date")
    d["Win_Rate"]       = (d["Winners"]/d["Trades"]*100).round(1)
    d["Gross_PnL"]      = d["Gross_PnL"].round(2)
    d["Cumulative_PnL"] = d["Gross_PnL"].cumsum().round(2)
    d["Running_Max"]    = d["Cumulative_PnL"].cummax().round(2)
    d["Drawdown"]       = (d["Cumulative_PnL"] - d["Running_Max"]).round(2)
    d["Avg_Duration"]   = d["Avg_Duration"].round(0).astype(int)
    return d


def build_monthly(t):
    m = t.groupby("ym").agg(
        Trades       =("pnl","count"),
        Winners      =("pnl", lambda x:(x>0).sum()),
        Gross_PnL    =("pnl","sum"),
        Avg_PnL      =("pnl","mean"),
        Best_Trade   =("pnl","max"),
        Worst_Trade  =("pnl","min"),
        Target_Exits =("result", lambda x:(x=="TARGET").sum()),
        SL_Exits     =("result", lambda x:(x=="SL").sum()),
        Time_Exits   =("result", lambda x:(x=="TIME").sum()),
        Avg_Duration =("duration_min","mean"),
    ).reset_index()
    m["Win_Rate"]       = (m["Winners"]/m["Trades"]*100).round(1)
    m["Gross_PnL"]      = m["Gross_PnL"].round(2)
    m["Avg_PnL"]        = m["Avg_PnL"].round(2)
    m["Cumulative_PnL"] = m["Gross_PnL"].cumsum().round(2)
    m["Avg_Duration"]   = m["Avg_Duration"].round(0)
    return m


def build_yearly(t):
    y = t.groupby("year").agg(
        Trades       =("pnl","count"),
        Winners      =("pnl", lambda x:(x>0).sum()),
        Gross_PnL    =("pnl","sum"),
        Best_Trade   =("pnl","max"),
        Worst_Trade  =("pnl","min"),
        Target_Exits =("result", lambda x:(x=="TARGET").sum()),
        SL_Exits     =("result", lambda x:(x=="SL").sum()),
        Time_Exits   =("result", lambda x:(x=="TIME").sum()),
        Avg_Duration =("duration_min","mean"),
    ).reset_index()
    y["Win_Rate"]       = (y["Winners"]/y["Trades"]*100).round(1)
    y["Gross_PnL"]      = y["Gross_PnL"].round(2)
    y["Cumulative_PnL"] = y["Gross_PnL"].cumsum().round(2)
    y["Avg_Duration"]   = y["Avg_Duration"].round(0)
    return y


def compute_metrics(t, daily):
    wins  = t[t["pnl"] > 0]["pnl"]
    loss  = t[t["pnl"] < 0]["pnl"]
    dr    = t.groupby("trade_date")["pnl"].sum()
    pf    = abs(wins.sum() / loss.sum())
    wlr   = wins.mean() / abs(loss.mean())
    sharpe  = dr.mean() / dr.std() * np.sqrt(252)
    sortino = dr.mean() / dr[dr < 0].std() * np.sqrt(252)
    max_dd  = daily["Drawdown"].min()
    total   = t["pnl"].sum()
    years   = (t["entry_ts"].max() - t["entry_ts"].min()).days / 365.25
    calmar  = (total / years) / abs(max_dd)

    # streaks
    loss_streak = win_streak = cur_l = cur_w = 0
    for p in t.sort_values(["trade_date","entry_time"])["pnl"]:
        if p < 0: cur_l += 1; loss_streak = max(loss_streak, cur_l); cur_w = 0
        else:     cur_w += 1; win_streak  = max(win_streak,  cur_w); cur_l = 0

    consec_loss_days = max(
        (len(list(g)) for k, g in itertools.groupby(daily["Gross_PnL"] < 0) if k),
        default=0)

    pctiles = {f"P{p}": round(float(np.percentile(t["pnl"], p)), 2)
               for p in [5, 10, 25, 50, 75, 90, 95, 99]}

    return {
        "Total Trades":           len(t),
        "Winners":                int((t["pnl"]>0).sum()),
        "Losers":                 int((t["pnl"]<0).sum()),
        "Win Rate %":             round((t["pnl"]>0).mean()*100, 2),
        "Total PnL":              round(total, 2),
        "Avg Win":                round(wins.mean(), 2),
        "Avg Loss":               round(loss.mean(), 2),
        "Largest Win":            round(t["pnl"].max(), 2),
        "Largest Loss":           round(t["pnl"].min(), 2),
        "Profit Factor":          round(pf, 3),
        "Win / Loss Ratio":       round(wlr, 3),
        "Expectancy (avg PnL)":   round(t["pnl"].mean(), 3),
        "Sharpe Ratio":           round(sharpe, 3),
        "Sortino Ratio":          round(sortino, 3),
        "Calmar Ratio":           round(calmar, 3),
        "Max Drawdown":           round(max_dd, 2),
        "Recovery Factor":        round(total / abs(max_dd), 3),
        "Max Losing Streak":      loss_streak,
        "Max Winning Streak":     win_streak,
        "Total Trading Days":     len(daily),
        "Profitable Days":        int((daily["Gross_PnL"]>0).sum()),
        "Day Win Rate %":         round((daily["Gross_PnL"]>0).mean()*100, 1),
        "Max Consec Loss Days":   consec_loss_days,
        "Avg PnL / Day":          round(dr.mean(), 2),
        "Daily PnL Std Dev":      round(dr.std(), 2),
        "Total Months":           len(t["ym"].unique()),
        "Profitable Months":      int((t.groupby("ym")["pnl"].sum() > 0).sum()),
        "Month Win Rate %":       round((t.groupby("ym")["pnl"].sum() > 0).mean()*100, 1),
        "TARGET Exits":           int((t["result"]=="TARGET").sum()),
        "SL Exits":               int((t["result"]=="SL").sum()),
        "TIME Exits":             int((t["result"]=="TIME").sum()),
        "Avg Duration (min)":     round(t["duration_min"].mean(), 1),
        "Median Duration (min)":  round(t["duration_min"].median(), 1),
        **pctiles,
    }


# ─── SHEET BUILDERS ───────────────────────────────────────────────────────────

def write_dashboard(wb, metrics, monthly, trades):
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False

    # Title
    title_row(ws, "NIFTY Options Strategy — Trade Log & Statistical Report", 8, row=1, sz=14, height=34)
    ws.merge_cells("A2:H2")
    c = ws.cell(2,1)
    c.value     = f"Period: {trades['trade_date'].min()}  →  {trades['trade_date'].max()}   |   {metrics['Total Trades']:,} trades across {metrics['Total Trading Days']:,} days"
    c.font      = font(sz=9, color="AAAAAA")
    c.fill      = fill(NAVY)
    c.alignment = C
    ws.row_dimensions[2].height = 16

    ws.row_dimensions[3].height = 10

    # KPI cards
    kpis = [
        ("Total Trades",      f"{metrics['Total Trades']:,}",     f"{metrics['Winners']} W / {metrics['Losers']} L",   "2563EB"),
        ("Win Rate",          f"{metrics['Win Rate %']}%",         f"{metrics['Profitable Days']} profitable days",     "16A34A"),
        ("Total PnL",         f"₹{metrics['Total PnL']:,.2f}",    f"₹{metrics['Avg PnL / Day']:+.2f} avg/day",         "1A7A3C" if metrics["Total PnL"]>0 else "C0392B"),
        ("Profit Factor",     f"{metrics['Profit Factor']}x",     f"Win/Loss ratio: {metrics['Win / Loss Ratio']}",    "7C3AED"),
        ("Sharpe Ratio",      f"{metrics['Sharpe Ratio']}",       f"Sortino: {metrics['Sortino Ratio']}",              "0E7490"),
        ("Max Drawdown",      f"₹{metrics['Max Drawdown']:,.2f}", f"Recovery factor: {metrics['Recovery Factor']}",    "C0392B"),
        ("Calmar Ratio",      f"{metrics['Calmar Ratio']}",       "Annual PnL ÷ Max DD",                               "EA8A23"),
        ("Month Win Rate",    f"{metrics['Month Win Rate %']}%",  f"{metrics['Profitable Months']}/{metrics['Total Months']} months",  "16A34A"),
    ]

    ws.merge_cells("A4:H4")
    ws.cell(4,1).value     = "KEY PERFORMANCE METRICS"
    ws.cell(4,1).font      = font(sz=8, color="AAAAAA")
    ws.cell(4,1).fill      = fill(LGRAY)
    ws.cell(4,1).alignment = L
    ws.row_dimensions[4].height = 16

    for i, (lbl, val, sub, col) in enumerate(kpis):
        c = i + 1
        for r in [5,6,7,8,9]:
            ws.cell(r,c).fill = fill(LGRAY); ws.cell(r,c).border = thin()
        ws.cell(5,c).value = lbl
        ws.cell(5,c).font  = font(sz=8, color="888888"); ws.cell(5,c).alignment = C
        ws.cell(6,c).value = val
        ws.cell(6,c).font  = font(sz=13, bold=True, color=col); ws.cell(6,c).alignment = C
        ws.cell(7,c).value = sub
        ws.cell(7,c).font  = font(sz=8, color="AAAAAA"); ws.cell(7,c).alignment = C
        ws.cell(8,c).value = ""; ws.cell(9,c).value = ""
    for r in [5,6,7,8,9]: ws.row_dimensions[r].height = 18
    ws.row_dimensions[9].height = 6

    # Second KPI block
    kpis2 = [
        ("Largest Win",        f"₹{metrics['Largest Win']:,.2f}",       "single trade",            "16A34A"),
        ("Largest Loss",       f"₹{metrics['Largest Loss']:,.2f}",      "single trade",            "C0392B"),
        ("Avg Win",            f"₹{metrics['Avg Win']:,.2f}",           "per winning trade",       "16A34A"),
        ("Avg Loss",           f"₹{metrics['Avg Loss']:,.2f}",          "per losing trade",        "C0392B"),
        ("Max Loss Streak",    f"{metrics['Max Losing Streak']} trades", "consecutive losses",     "C0392B"),
        ("Max Win Streak",     f"{metrics['Max Winning Streak']} trades","consecutive wins",       "16A34A"),
        ("Avg Duration",       f"{metrics['Avg Duration (min)']} min",  f"median {metrics['Median Duration (min)']} min","2563EB"),
        ("Consec Loss Days",   f"{metrics['Max Consec Loss Days']} days","max losing days in row", "EA8A23"),
    ]
    for i, (lbl, val, sub, col) in enumerate(kpis2):
        c = i + 1
        for r in [10,11,12,13,14]:
            ws.cell(r,c).fill = fill(WHITE); ws.cell(r,c).border = thin()
        ws.cell(10,c).value = lbl
        ws.cell(10,c).font  = font(sz=8, color="888888"); ws.cell(10,c).alignment = C
        ws.cell(11,c).value = val
        ws.cell(11,c).font  = font(sz=11, bold=True, color=col); ws.cell(11,c).alignment = C
        ws.cell(12,c).value = sub
        ws.cell(12,c).font  = font(sz=8, color="AAAAAA"); ws.cell(12,c).alignment = C
        ws.cell(13,c).value = ""; ws.cell(14,c).value = ""
    for r in [10,11,12,13,14]: ws.row_dimensions[r].height = 17
    ws.row_dimensions[14].height = 8

    # Monthly table
    ws.merge_cells("A15:H15")
    ws.cell(15,1).value     = "MONTHLY PERFORMANCE SUMMARY"
    ws.cell(15,1).font      = font(sz=10, bold=True, color="FFFFFF")
    ws.cell(15,1).fill      = fill(NAVY)
    ws.cell(15,1).alignment = L
    ws.cell(15,1).border    = thin()
    ws.row_dimensions[15].height = 22

    mhdr = ["Month","Trades","Winners","Win Rate","Targets","SL Exits","Time Exits","Gross PnL","Cumulative PnL"]
    header_row(ws, 16, mhdr, height=18)

    for i, (_, row) in enumerate(monthly.iterrows()):
        r   = 17 + i
        pnl = row["Gross_PnL"]
        cum = row["Cumulative_PnL"]
        bg  = GREEN if pnl > 0 else RED
        vals = [row["ym"], row["Trades"], row["Winners"], row["Win_Rate"],
                row["Target_Exits"], row["SL_Exits"], row["Time_Exits"], pnl, cum]
        for c, v in enumerate(vals, 1):
            fnt = pnl_font(v, sz=9, bold=c==9) if c in (8,9) else font(sz=9)
            fmt = "#,##0.00" if c in (8,9) else ("0.0\"%\"" if c==4 else None)
            wcell(ws, r, c, v, bg=bg, fnt=fnt, fmt=fmt)
        ws.row_dimensions[r].height = 15

    col_widths(ws, [14,9,9,10,10,9,11,14,16])


def write_trade_log(wb, trades):
    ws = wb.create_sheet("Daily Trade Log")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    title_row(ws, "Daily Trade Log  —  Every Individual Trade", 14, row=1, height=26)

    hdrs = ["Date","Entry Time","Exit Time","Ticker","Type","Strike",
            "Entry Price","SL","Target","Exit Price","PnL","PnL %",
            "Result","Duration (min)"]
    header_row(ws, 2, hdrs, height=20)

    RCOL = {"TARGET":"1A7A3C","SL":"C0392B","TIME":"2563EB"}

    for i, (_, row) in enumerate(trades.iterrows()):
        r   = 3 + i
        pnl = row["pnl"]
        bg  = GREEN if pnl > 0 else (RED if pnl < 0 else WHITE)
        vals = [str(row["trade_date"]), row["entry_time"], row["exit_time"],
                row["Ticker"], row["option_type"],
                int(row["strike"]) if pd.notna(row["strike"]) else "",
                row["entry_price"],
                row["sl"]     if pd.notna(row["sl"])     else "",
                row["target"] if pd.notna(row["target"]) else "",
                row["exit_price"], pnl, row["pnl_pct"], row["result"], row["duration_min"]]
        for c, v in enumerate(vals, 1):
            if c == 11:   fnt = pnl_font(pnl, sz=9, bold=True); fmt = "#,##0.00"
            elif c == 12: fnt = pnl_font(pnl, sz=9);            fmt = "0.00\"%\""
            elif c == 13: fnt = font(sz=9, bold=True, color=RCOL.get(str(v),"000000")); fmt = None
            else:         fnt = font(sz=9); fmt = None
            wcell(ws, r, c, v, bg=bg, fnt=fnt, fmt=fmt)
        ws.row_dimensions[r].height = 15

    # totals
    tr = 3 + len(trades)
    totals_row(ws, tr, 14, {11: f"=SUM(K3:K{tr-1})"})

    col_widths(ws, [12,10,10,26,6,8,10,8,8,10,10,8,9,14])


def write_daily_summary(wb, daily):
    ws = wb.create_sheet("Daily Summary")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    title_row(ws, "Daily Summary  —  PnL, Win Rate & Exit Breakdown per Day", 13, height=26)

    hdrs = ["Date","Trades","Winners","Losers","Win Rate %",
            "Target","SL","Time","Gross PnL","Best Trade","Worst Trade",
            "Avg Duration","Cumulative PnL"]
    header_row(ws, 2, hdrs, height=20)

    for i, (_, row) in enumerate(daily.iterrows()):
        r   = 3 + i
        pnl = row["Gross_PnL"]
        bg  = GREEN if pnl > 0 else (RED if pnl < 0 else WHITE)
        vals = [str(row["trade_date"]), row["Trades"], row["Winners"], row["Losers"],
                row["Win_Rate"], row["Target_Count"], row["SL_Count"], row["Time_Count"],
                pnl, row["Best_Trade"], row["Worst_Trade"], row["Avg_Duration"], row["Cumulative_PnL"]]
        for c, v in enumerate(vals, 1):
            if c == 5:  fnt = font(sz=9); fmt = "0.0\"%\""
            elif c in (9,10,11,13): fnt = pnl_font(v, sz=9, bold=c==13); fmt = "#,##0.00"
            else:       fnt = font(sz=9); fmt = None
            wcell(ws, r, c, v, bg=bg, fnt=fnt, fmt=fmt)
        ws.row_dimensions[r].height = 15

    tr = 3 + len(daily)
    totals_row(ws, tr, 13, {9: f"=SUM(I3:I{tr-1})", 13: f"=I{tr}"})
    col_widths(ws, [13,7,7,7,10,7,6,6,12,12,12,12,14])


def write_monthly(wb, monthly):
    ws = wb.create_sheet("Monthly Summary")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    title_row(ws, "Monthly Summary  —  Strategy Performance by Calendar Month", 11, height=26)

    hdrs = ["Month","Trades","Winners","Win Rate %","Targets","SL Exits","Time Exits",
            "Gross PnL","Avg PnL","Best Trade","Cumulative PnL"]
    header_row(ws, 2, hdrs, height=20)

    for i, (_, row) in enumerate(monthly.iterrows()):
        r   = 3 + i
        pnl = row["Gross_PnL"]
        bg  = GREEN if pnl > 0 else RED
        vals = [row["ym"], row["Trades"], row["Winners"], row["Win_Rate"],
                row["Target_Exits"], row["SL_Exits"], row["Time_Exits"],
                pnl, row["Avg_PnL"], row["Best_Trade"], row["Cumulative_PnL"]]
        for c, v in enumerate(vals, 1):
            if c == 4:  fnt = font(sz=10); fmt = "0.0\"%\""
            elif c in (8,9,10,11): fnt = pnl_font(v, sz=10, bold=c in (8,11)); fmt = "#,##0.00"
            else:       fnt = font(sz=10); fmt = None
            wcell(ws, r, c, v, bg=bg, fnt=fnt, fmt=fmt)
        ws.row_dimensions[r].height = 16

    tr = 3 + len(monthly)
    totals_row(ws, tr, 11,
               {8: f"=SUM(H3:H{tr-1})", 9: f"=AVERAGE(I3:I{tr-1})", 11: f"=H{tr}"})
    col_widths(ws, [14,8,8,10,9,9,11,14,12,13,16])


def write_yearly(wb, yearly):
    ws = wb.create_sheet("Yearly Summary")
    ws.sheet_view.showGridLines = False

    title_row(ws, "Yearly Summary  —  Annual Performance Breakdown", 10, height=26)

    hdrs = ["Year","Trades","Winners","Win Rate %","Targets","SL Exits","Time Exits",
            "Gross PnL","Best Trade","Worst Trade","Cumulative PnL"]
    header_row(ws, 2, hdrs, height=20)

    for i, (_, row) in enumerate(yearly.iterrows()):
        r   = 3 + i
        pnl = row["Gross_PnL"]
        bg  = GREEN if pnl > 0 else RED
        vals = [row["year"], row["Trades"], row["Winners"], row["Win_Rate"],
                row["Target_Exits"], row["SL_Exits"], row["Time_Exits"],
                pnl, row["Best_Trade"], row["Worst_Trade"], row["Cumulative_PnL"]]
        for c, v in enumerate(vals, 1):
            if c == 4:  fnt = font(sz=10); fmt = "0.0\"%\""
            elif c in (8,9,10,11): fnt = pnl_font(v, sz=10, bold=c in (8,11)); fmt = "#,##0.00"
            else:       fnt = font(sz=10); fmt = None
            wcell(ws, r, c, v, bg=bg, fnt=fnt, fmt=fmt)
        ws.row_dimensions[r].height = 18

    tr = 3 + len(yearly)
    totals_row(ws, tr, 11,
               {8: f"=SUM(H3:H{tr-1})", 11: f"=H{tr}"})
    col_widths(ws, [8,8,8,10,9,9,11,14,13,13,14])


def write_statistics(wb, metrics, trades, pctiles=None):
    ws = wb.create_sheet("Statistics")
    ws.sheet_view.showGridLines = False

    title_row(ws, "Full Statistical Report  —  All Performance Metrics", 4, height=28)

    sections = [
        ("CORE PERFORMANCE", [
            ("Total Trades",            metrics["Total Trades"]),
            ("Winners",                 metrics["Winners"]),
            ("Losers",                  metrics["Losers"]),
            ("Win Rate %",              f"{metrics['Win Rate %']}%"),
            ("Total PnL",               f"₹{metrics['Total PnL']:,.2f}"),
            ("Avg PnL / Trade",         f"₹{metrics['Expectancy (avg PnL)']:.3f}"),
            ("Avg PnL / Day",           f"₹{metrics['Avg PnL / Day']:.2f}"),
            ("Daily PnL Std Dev",       f"₹{metrics['Daily PnL Std Dev']:.2f}"),
        ]),
        ("TRADE QUALITY", [
            ("Avg Win",                 f"₹{metrics['Avg Win']:.2f}"),
            ("Avg Loss",                f"₹{metrics['Avg Loss']:.2f}"),
            ("Largest Win",             f"₹{metrics['Largest Win']:.2f}"),
            ("Largest Loss",            f"₹{metrics['Largest Loss']:.2f}"),
            ("Win / Loss Ratio",        metrics["Win / Loss Ratio"]),
            ("Profit Factor",           metrics["Profit Factor"]),
            ("Expectancy",              f"₹{metrics['Expectancy (avg PnL)']:.3f}"),
        ]),
        ("RISK METRICS", [
            ("Sharpe Ratio (ann.)",     metrics["Sharpe Ratio"]),
            ("Sortino Ratio (ann.)",    metrics["Sortino Ratio"]),
            ("Calmar Ratio",            metrics["Calmar Ratio"]),
            ("Max Drawdown",            f"₹{metrics['Max Drawdown']:.2f}"),
            ("Recovery Factor",         metrics["Recovery Factor"]),
            ("Max Losing Streak",       f"{metrics['Max Losing Streak']} trades"),
            ("Max Winning Streak",      f"{metrics['Max Winning Streak']} trades"),
            ("Max Consec Loss Days",    f"{metrics['Max Consec Loss Days']} days"),
        ]),
        ("DAY / MONTH STATS", [
            ("Total Trading Days",      metrics["Total Trading Days"]),
            ("Profitable Days",         metrics["Profitable Days"]),
            ("Day Win Rate %",          f"{metrics['Day Win Rate %']}%"),
            ("Total Months",            metrics["Total Months"]),
            ("Profitable Months",       metrics["Profitable Months"]),
            ("Month Win Rate %",        f"{metrics['Month Win Rate %']}%"),
        ]),
        ("EXIT BREAKDOWN", [
            ("TARGET Exits",            metrics["TARGET Exits"]),
            ("SL Exits",                metrics["SL Exits"]),
            ("TIME Exits",              metrics["TIME Exits"]),
            ("Avg Duration (min)",      metrics["Avg Duration (min)"]),
            ("Median Duration (min)",   metrics["Median Duration (min)"]),
        ]),
        ("PnL PERCENTILES", [
            ("P5  (5th percentile)",    f"₹{metrics['P5']:.2f}"),
            ("P10 (10th percentile)",   f"₹{metrics['P10']:.2f}"),
            ("P25 (25th percentile)",   f"₹{metrics['P25']:.2f}"),
            ("P50 (median)",            f"₹{metrics['P50']:.2f}"),
            ("P75 (75th percentile)",   f"₹{metrics['P75']:.2f}"),
            ("P90 (90th percentile)",   f"₹{metrics['P90']:.2f}"),
            ("P95 (95th percentile)",   f"₹{metrics['P95']:.2f}"),
            ("P99 (99th percentile)",   f"₹{metrics['P99']:.2f}"),
        ]),
    ]

    row = 2
    for sec_title, rows in sections:
        ws.row_dimensions[row].height = 8
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        c = ws.cell(row, 1)
        c.value = sec_title; c.font = font(sz=9, bold=True, color="FFFFFF")
        c.fill = fill(HEADER); c.alignment = L; c.border = thin()
        ws.row_dimensions[row].height = 20
        row += 1

        for lbl, val in rows:
            wcell(ws, row, 1, lbl, bg=LGRAY, fnt=font(sz=10, color="444444"), align=L)
            wcell(ws, row, 2, val, bg=WHITE, fnt=font(sz=10, bold=True))
            wcell(ws, row, 3, "", bg=WHITE, border=False)
            wcell(ws, row, 4, "", bg=WHITE, border=False)
            ws.row_dimensions[row].height = 18
            row += 1

    col_widths(ws, [30, 22, 4, 4])


def write_breakdown(wb, trades):
    ws = wb.create_sheet("Breakdown Analysis")
    ws.sheet_view.showGridLines = False

    title_row(ws, "Breakdown Analysis  —  CE/PE · Day of Week · Hour · Duration · Entry Price", 14, height=26)

    # Helper to write a sub-table
    def sub_table(ws, start_row, start_col, title, headers, rows, widths=None, ncols=None):
        nc = ncols or len(headers)
        ws.merge_cells(start_row=start_row, start_column=start_col,
                       end_row=start_row, end_column=start_col+nc-1)
        c = ws.cell(start_row, start_col)
        c.value = title; c.font = font(sz=10, bold=True, color="FFFFFF")
        c.fill = fill(NAVY); c.alignment = L; c.border = thin()
        ws.row_dimensions[start_row].height = 20

        for ci, h in enumerate(headers):
            cell = ws.cell(start_row+1, start_col+ci)
            cell.value = h; cell.font = font(sz=9, bold=True, color="FFFFFF")
            cell.fill = fill(HEADER); cell.alignment = C; cell.border = thin()
        ws.row_dimensions[start_row+1].height = 18

        for ri, row_vals in enumerate(rows):
            r = start_row + 2 + ri
            for ci, v in enumerate(row_vals):
                bg = GREEN if (isinstance(v, float) and ci > 1 and v > 0) else \
                     (RED if (isinstance(v, float) and ci > 1 and v < 0) else \
                     (ALT if ri % 2 else WHITE))
                fnt = pnl_font(v, sz=9) if (isinstance(v, float) and "₹" not in str(v) and ci > 2) else font(sz=9)
                wcell(ws, r, start_col+ci, v, bg=bg, fnt=fnt)
            ws.row_dimensions[r].height = 15

    # 1. CE vs PE
    ce_pe = trades.groupby("option_type").agg(
        Trades=("pnl","count"), Winners=("pnl", lambda x:(x>0).sum()),
        Total_PnL=("pnl","sum"), Avg_PnL=("pnl","mean"),
        Targets=("result", lambda x:(x=="TARGET").sum()),
        SLs=("result", lambda x:(x=="SL").sum()),
    ).reset_index()
    ce_pe["Win_Rate"] = (ce_pe["Winners"]/ce_pe["Trades"]*100).round(1)
    sub_table(ws, 2, 1, "CE vs PE Performance",
        ["Type","Trades","Win Rate","Total PnL","Avg PnL","Targets","SLs"],
        [[r["option_type"], r["Trades"], f"{r['Win_Rate']}%",
          round(r["Total_PnL"],2), round(r["Avg_PnL"],2), r["Targets"], r["SLs"]]
         for _, r in ce_pe.iterrows()])

    # 2. Day of week
    dow_map = {"Monday":0,"Tuesday":1,"Wednesday":2,"Thursday":3,"Friday":4}
    dow = trades.groupby("dow_name").agg(
        Trades=("pnl","count"), Winners=("pnl", lambda x:(x>0).sum()),
        Total_PnL=("pnl","sum"), Avg_PnL=("pnl","mean"),
        Targets=("result", lambda x:(x=="TARGET").sum()),
        SLs=("result", lambda x:(x=="SL").sum()),
        Avg_Duration=("duration_min","mean"),
    ).reset_index().sort_values("dow_name", key=lambda x: x.map(dow_map))
    dow["Win_Rate"] = (dow["Winners"]/dow["Trades"]*100).round(1)
    sub_table(ws, 7, 1, "Day of Week Performance",
        ["Day","Trades","Win Rate","Total PnL","Avg PnL","Targets","SLs","Avg Dur (min)"],
        [[r["dow_name"], r["Trades"], f"{r['Win_Rate']}%",
          round(r["Total_PnL"],2), round(r["Avg_PnL"],2), r["Targets"], r["SLs"],
          round(r["Avg_Duration"],0)] for _, r in dow.iterrows()])

    # 3. Hour of entry
    hr = trades.groupby("entry_hour").agg(
        Trades=("pnl","count"), Winners=("pnl", lambda x:(x>0).sum()),
        Total_PnL=("pnl","sum"), Avg_PnL=("pnl","mean"),
        SLs=("result", lambda x:(x=="SL").sum()),
    ).reset_index()
    hr["Win_Rate"] = (hr["Winners"]/hr["Trades"]*100).round(1)
    hr["SL_Rate"]  = (hr["SLs"]/hr["Trades"]*100).round(1)
    sub_table(ws, 20, 1, "Entry Hour Performance",
        ["Hour","Trades","Win Rate","Total PnL","Avg PnL","SL Rate"],
        [[f"{int(r['entry_hour']):02d}:xx", r["Trades"], f"{r['Win_Rate']}%",
          round(r["Total_PnL"],2), round(r["Avg_PnL"],2), f"{r['SL_Rate']}%"]
         for _, r in hr.iterrows()])

    # 4. Duration bucket
    dur = trades.groupby("dur_bucket", observed=True).agg(
        Trades=("pnl","count"), Winners=("pnl", lambda x:(x>0).sum()),
        Total_PnL=("pnl","sum"), Avg_PnL=("pnl","mean"),
    ).reset_index()
    dur["Win_Rate"] = (dur["Winners"]/dur["Trades"]*100).round(1)
    sub_table(ws, 2, 10, "Duration Bucket Performance",
        ["Duration","Trades","Win Rate","Total PnL","Avg PnL"],
        [[str(r["dur_bucket"]), r["Trades"], f"{r['Win_Rate']}%",
          round(r["Total_PnL"],2), round(r["Avg_PnL"],2)]
         for _, r in dur.iterrows()])

    # 5. Entry price bucket
    ep = trades.groupby("ep_bucket", observed=True).agg(
        Trades=("pnl","count"), Winners=("pnl", lambda x:(x>0).sum()),
        Total_PnL=("pnl","sum"), Avg_PnL=("pnl","mean"),
        SLs=("result", lambda x:(x=="SL").sum()),
    ).reset_index()
    ep["Win_Rate"] = (ep["Winners"]/ep["Trades"]*100).round(1)
    ep["SL_Rate"]  = (ep["SLs"]/ep["Trades"]*100).round(1)
    sub_table(ws, 13, 10, "Entry Price Bucket",
        ["Entry Price","Trades","Win Rate","Total PnL","Avg PnL","SL Rate"],
        [[str(r["ep_bucket"]), r["Trades"], f"{r['Win_Rate']}%",
          round(r["Total_PnL"],2), round(r["Avg_PnL"],2), f"{r['SL_Rate']}%"]
         for _, r in ep.iterrows()])

    # 6. Result type stats
    res = trades.groupby("result").agg(
        Trades=("pnl","count"), Total_PnL=("pnl","sum"), Avg_PnL=("pnl","mean"),
        Best=("pnl","max"), Worst=("pnl","min"), Avg_Duration=("duration_min","mean"),
    ).reset_index()
    sub_table(ws, 23, 10, "Exit Type Statistics",
        ["Result","Trades","Total PnL","Avg PnL","Best","Worst","Avg Dur (min)"],
        [[r["result"], r["Trades"], round(r["Total_PnL"],2), round(r["Avg_PnL"],2),
          round(r["Best"],2), round(r["Worst"],2), round(r["Avg_Duration"],0)]
         for _, r in res.iterrows()])

    col_widths(ws, [12,8,9,12,10,8,8,12, 4, 12,8,9,12,10,8,8])


def write_best_worst(wb, daily, trades):
    ws = wb.create_sheet("Best & Worst")
    ws.sheet_view.showGridLines = False

    title_row(ws, "Best & Worst Days / Months / Trades", 8, height=26)

    def mini_table(ws, r, c, title, hdrs, rows, nc):
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+nc-1)
        cell = ws.cell(r, c)
        cell.value = title; cell.font = font(sz=10, bold=True, color="FFFFFF")
        cell.fill = fill(NAVY); cell.alignment = L; cell.border = thin()
        ws.row_dimensions[r].height = 20
        for ci, h in enumerate(hdrs):
            cell = ws.cell(r+1, c+ci)
            cell.value = h; cell.font = font(sz=9, bold=True, color="FFFFFF")
            cell.fill = fill(HEADER); cell.alignment = C; cell.border = thin()
        ws.row_dimensions[r+1].height = 18
        for ri, rv in enumerate(rows):
            for ci, v in enumerate(rv):
                rr = r + 2 + ri
                bg = GREEN if ri < len(rows)//2 else RED
                wcell(ws, rr, c+ci, v, bg=bg, fnt=pnl_font(v, sz=9) if isinstance(v,float) else font(sz=9))
                ws.row_dimensions[rr].height = 16

    # Best 10 days
    b10 = daily.sort_values("Gross_PnL", ascending=False).head(10)
    mini_table(ws, 2, 1, "Top 10 Best Days", ["Date","Gross PnL","Trades","Win Rate"],
               [[str(r["trade_date"]), r["Gross_PnL"], r["Trades"], f"{r['Win_Rate']}%"]
                for _, r in b10.iterrows()], 4)

    # Worst 10 days
    w10 = daily.sort_values("Gross_PnL").head(10)
    mini_table(ws, 2, 6, "Top 10 Worst Days", ["Date","Gross PnL","Trades","Win Rate"],
               [[str(r["trade_date"]), r["Gross_PnL"], r["Trades"], f"{r['Win_Rate']}%"]
                for _, r in w10.iterrows()], 4)

    # Best 10 individual trades
    bt = trades.sort_values("pnl", ascending=False).head(10)
    mini_table(ws, 16, 1, "Top 10 Best Trades",
               ["Date","Ticker","Type","Entry","Exit","PnL","Result"],
               [[str(r["trade_date"]), r["Ticker"], r["option_type"],
                 r["entry_price"], r["exit_price"], r["pnl"], r["result"]]
                for _, r in bt.iterrows()], 7)

    # Worst 10 individual trades
    wt = trades.sort_values("pnl").head(10)
    mini_table(ws, 16, 9, "Top 10 Worst Trades",
               ["Date","Ticker","Type","Entry","Exit","PnL","Result"],
               [[str(r["trade_date"]), r["Ticker"], r["option_type"],
                 r["entry_price"], r["exit_price"], r["pnl"], r["result"]]
                for _, r in wt.iterrows()], 7)

    col_widths(ws, [12,12,6,8,8,10,9, 4, 12,26,6,8,8,10,9])


def write_drawdown(wb, daily):
    ws = wb.create_sheet("Drawdown")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    title_row(ws, "Drawdown Analysis  —  Daily Cumulative PnL & Drawdown", 6, height=26)

    hdrs = ["Date","Gross PnL","Cumulative PnL","Running Max","Drawdown","DD %"]
    header_row(ws, 2, hdrs, height=20)

    for i, (_, row) in enumerate(daily.iterrows()):
        r    = 3 + i
        pnl  = row["Gross_PnL"]
        cum  = row["Cumulative_PnL"]
        dd   = row["Drawdown"]
        rmax = row["Running_Max"]
        ddpct = round(dd / rmax * 100, 2) if rmax and rmax != 0 else 0.0
        bg   = GREEN if pnl > 0 else (RED if pnl < 0 else WHITE)
        vals = [str(row["trade_date"]), pnl, cum, rmax, dd, ddpct]
        for c, v in enumerate(vals, 1):
            if c in (2,3,4,5): fnt = pnl_font(v, sz=9); fmt = "#,##0.00"
            elif c == 6:       fnt = pnl_font(v, sz=9); fmt = "0.00\"%\""
            else:              fnt = font(sz=9);         fmt = None
            wcell(ws, r, c, v, bg=bg, fnt=fnt, fmt=fmt)
        ws.row_dimensions[r].height = 15

    col_widths(ws, [13,13,14,13,13,10])


# ─── MAIN ─────────────────────────────────────────────────────────────────────
def generate_trade_logs(INPUT_PATH):
    print("Reading order book...")
    strategy_save_dir = INPUT_PATH
    file_path = strategy_save_dir /"order_book.csv"
    trades  = build_trades(file_path)
    daily   = build_daily(trades)
    monthly = build_monthly(trades)
    yearly  = build_yearly(trades)
    metrics = compute_metrics(trades, daily)

    print(f"  {metrics['Total Trades']:,} trades | "
          f"{metrics['Total Trading Days']:,} days | "
          f"Win rate {metrics['Win Rate %']}% | "
          f"Total PnL ₹{metrics['Total PnL']:,.2f}")

    print("Building workbook...")
    wb = Workbook()

    write_dashboard(wb, metrics, monthly, trades)
    write_trade_log(wb, trades)
    write_daily_summary(wb, daily)
    write_monthly(wb, monthly)
    write_yearly(wb, yearly)
    write_statistics(wb, metrics, trades)
    write_breakdown(wb, trades)
    write_best_worst(wb, daily, trades)
    write_drawdown(wb, daily)

    wb.save(strategy_save_dir / OUTPUT_PATH)
    print(f"Saved → {OUTPUT_PATH}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    generate_trade_logs()
